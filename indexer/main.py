import os
import sys
import argparse
import sqlite3
import json
import time
import base64
import mimetypes
import logging
from typing import List, Dict, Any, Optional

# Third-party imports
import requests
import pandas as pd
from openpyxl import load_workbook
import docx
import PyPDF2

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

if not os.path.exists("./files"):
    os.mkdir("./files")

if not os.path.exists("./static"):
    os.mkdir("./static")

# --- Constants & Configuration ---
ALLOWED_TEXT_EXTENSIONS = {
    '.txt', '.md', '.java', '.py', '.c', '.cpp', '.h', '.html', '.css', '.js', 
    '.json', '.xml', '.yaml', '.yml', '.sql', '.sh', '.bat', '.csv'
}
ALLOWED_DOC_EXTENSIONS = {'.docx', '.doc'}
ALLOWED_XLS_EXTENSIONS = {'.xlsx', '.xls'}
ALLOWED_PDF_EXTENSIONS = {'.pdf'}
ALLOWED_IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.bmp', '.webp'}

# --- Database Management ---
def init_db(db_path: str):
    """Initializes the database with the new schema."""
    # Ensure directory exists
    db_dir = os.path.dirname(db_path)
    if db_dir and not os.path.exists(db_dir):
        try:
            os.makedirs(db_dir)
            logger.info(f"Created database directory: {db_dir}")
        except OSError as e:
            logger.error(f"Failed to create database directory {db_dir}: {e}")
            raise

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Check if table exists
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='FILE'")
    if not cursor.fetchone():
        # Create new table if not exists with all columns
        cursor.execute("""
            CREATE TABLE FILE (
                FileHash TEXT PRIMARY KEY,
                FileName TEXT UNIQUE NOT NULL,
                FileType TEXT NOT NULL,
                FileSize INTEGER NOT NULL,
                Time INTEGER NOT NULL,
                Description TEXT,
                ShortDescription TEXT,
                Subject TEXT,
                Year TEXT,
                Keywords TEXT,
                Embedding TEXT
            )
        """)
    else:
        # Migrate existing table: Add columns if they don't exist
        existing_cols = set()
        cursor.execute("PRAGMA table_info(FILE)")
        for col in cursor.fetchall():
            existing_cols.add(col[1])
        
        new_cols = {
            "Description": "TEXT",
            "ShortDescription": "TEXT",
            "Subject": "TEXT",
            "Year": "TEXT",
            "Keywords": "TEXT",
            "Embedding": "TEXT" 
        }
        
        for col_name, col_type in new_cols.items():
            if col_name not in existing_cols:
                logger.info(f"Adding missing column: {col_name}")
                try:
                    cursor.execute(f"ALTER TABLE FILE ADD COLUMN {col_name} {col_type}")
                except Exception as e:
                    logger.error(f"Failed to add column {col_name}: {e}")

    # Check/Create LOG table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS INDEX_LOG (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            StartTime TEXT,
            EndTime TEXT,
            TotalFiles INTEGER DEFAULT 0,
            SuccessCount INTEGER DEFAULT 0,
            FailedCount INTEGER DEFAULT 0,
            SkippedCount INTEGER DEFAULT 0,
            TypeStats TEXT
        )
    """)

    conn.commit()
    conn.close()

class IndexSession:
    def __init__(self, db_path):
        self.db_path = db_path
        self.start_time = time.strftime('%Y-%m-%d %H:%M:%S')
        self.total = 0
        self.success = 0
        self.failed = 0
        self.skipped = 0
        self.type_stats = {}
        self.session_id = self._create_session()
        
    def _create_session(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("INSERT INTO INDEX_LOG (StartTime) VALUES (?)", (self.start_time,))
        session_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return session_id

    def update(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE INDEX_LOG 
            SET TotalFiles=?, SuccessCount=?, FailedCount=?, SkippedCount=?, TypeStats=?
            WHERE id=?
        """, (self.total, self.success, self.failed, self.skipped, json.dumps(self.type_stats), self.session_id))
        conn.commit()
        conn.close()
        
    def finish(self):
        end_time = time.strftime('%Y-%m-%d %H:%M:%S')
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE INDEX_LOG 
            SET EndTime=?, TotalFiles=?, SuccessCount=?, FailedCount=?, SkippedCount=?, TypeStats=?
            WHERE id=?
        """, (end_time, self.total, self.success, self.failed, self.skipped, json.dumps(self.type_stats), self.session_id))
        conn.commit()
        conn.close()

    def log_success(self, file_type):
        self.total += 1
        self.success += 1
        self.type_stats[file_type] = self.type_stats.get(file_type, 0) + 1
        self.update()

    def log_fail(self, file_type):
        self.total += 1
        self.failed += 1
        self.type_stats[file_type] = self.type_stats.get(file_type, 0) + 1
        self.update()

    def log_skip(self, file_type):
        self.total += 1
        self.skipped += 1
        self.type_stats[file_type] = self.type_stats.get(file_type, 0) + 1
        # self.update() # Optional: limit validation frequency

def save_to_db(db_path: str, data: Dict[str, Any]):
    """Saves or updates file info in the database."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # We use INSERT OR REPLACE. 
    # Note: This might overwrite existing manual edits if any.
    cursor.execute("""
        INSERT OR REPLACE INTO FILE 
        (FileHash, FileName, FileType, FileSize, Time, Description, ShortDescription, Subject, Year, Keywords, Embedding)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data['FileHash'],
        data['FileName'],
        data['FileType'],
        data['FileSize'],
        data['Time'],
        data['Description'],
        data['ShortDescription'],
        data['Subject'],
        data['Year'],
        data['Keywords'],
        json.dumps(data['Embedding']) if data['Embedding'] else None
    ))
    
    conn.commit()
    conn.close()

# --- File Processing ---

def extract_text_from_file(file_path: str, ext: str) -> Optional[str]:
    """Extracts text from various file formats."""
    try:
        if ext in ALLOWED_TEXT_EXTENSIONS:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()[:5000] # Limit context
        
        elif ext in ALLOWED_DOC_EXTENSIONS:
            if ext == '.docx':
                doc = docx.Document(file_path)
                return "\n".join([para.text for para in doc.paragraphs])[:5000]
            # .doc support is tricky without system libraries, skipping for now
            
        elif ext in ALLOWED_XLS_EXTENSIONS:
            if ext == '.xlsx':
                wb = load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                data = []
                for row in ws.iter_rows(max_row=50): # Limit rows
                    row_data = [str(cell.value) for cell in row if cell.value]
                    data.append(", ".join(row_data))
                return "\n".join(data)
            
        elif ext in ALLOWED_PDF_EXTENSIONS:
            text = ""
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                # Read first 5 pages max
                for i in range(min(len(reader.pages), 5)):
                    page = reader.pages[i]
                    text += page.extract_text() or ""
            return text[:5000]
            
    except Exception as e:
        logger.warning(f"Error extracting text from {file_path}: {e}")
    
    return None

def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')

def check_if_indexed(db_path: str, file_hash: str) -> bool:
    """Checks if a file hash already exists in the database."""
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT 1 FROM FILE WHERE FileHash = ?", (file_hash,))
        exists = cursor.fetchone() is not None
        conn.close()
        return exists
    except Exception as e:
        logger.error(f"DB Check Error: {e}")
        return False

# --- AI Interaction ---

class AIClient:
    def __init__(self, host, key, model):
        self.host = host if host.endswith('/v1') else host + '/v1'
        self.key = key
        self.model = model
    
    def chat_completion(self, messages, json_mode=False):
        headers = {
            "Authorization": f"Bearer {self.key}",
            "Content-Type": "application/json"
        }
        data = {
            "model": self.model,
            "messages": messages,
            "response_format": {"type": "json_object"} if json_mode else None
        }
        try:
            response = requests.post(f"{self.host}/chat/completions", headers=headers, json=data, timeout=60)
            response.raise_for_status()
            return response.json()['choices'][0]['message']['content']
        except Exception as e:
            logger.error(f"AI API Error: {e}")
            return None

    def get_embedding(self, text):
        if hasattr(self, 'local_model') and self.local_model:
            # Local Mode
            return self.local_model.encode(text).tolist()

        headers = {
            "Authorization": f"Bearer {self.key}",
            "Content-Type": "application/json"
        }
        data = {
            "model": self.model, # "text-embedding-3-small" or similar
            "input": text
        }
        try:
            response = requests.post(f"{self.host}/embeddings", headers=headers, json=data, timeout=30)
            response.raise_for_status()
            return response.json()['data'][0]['embedding']
        except Exception as e:
            logger.warning(f"Embedding API Error: {e}")
            return None

def analyze_content(ai1: AIClient, ai2: AIClient, file_path: str, meta: Dict, context: str, file_text: Optional[str], embedding_client: Optional[AIClient] = None) -> Dict:
    """
    Orchestrates the analysis:
    1. If image/scanned PDF -> Use AI2 to get text.
    2. Analyze text/metadata with AI1.
    """
    
    # Step 1: Get Content
    content_to_analyze = file_text
    
    ext = meta['FileType_Ext']
    if not content_to_analyze and ext in ALLOWED_IMAGE_EXTENSIONS:
        # Use AI2 (OCR/Vision)
        logger.info("Using AI2 for OCR...")
        base64_image = encode_image(file_path)
        
        # Determine mime type
        mime_type = "image/jpeg"
        if ext.lower() == ".png":
            mime_type = "image/png"
        elif ext.lower() == ".webp":
            mime_type = "image/webp"
            
        messages = [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": "Extract all text from this image directly. Do not describe the image, just output the text content found inside."},
                    {
                        "type": "image_url", 
                        "image_url": {
                            "url": f"data:{mime_type};base64,{base64_image}"
                        }
                    }
                ]
            }
        ]
        content_to_analyze = ai2.chat_completion(messages)
    
    if not content_to_analyze:
        content_to_analyze = "[Binary or unreadable file content]"

    # Step 2: Analyze with AI1
    system_prompt = """
    You are a professional file archivist. Your job is to analyze file metadata, context, and snippets to generate structured metadata.
    Output JSON with the following keys:
    - description: A detailed description of what this file contains (max 200 words). MUST BE IN CHINESE (简体中文).
    - short_description: A very brief summary (max 30 words). MUST BE IN CHINESE (简体中文).
    - subject: The academic subject associated with this file (e.g., "Linear Algebra", "History"). Infer from filename and path. MUST BE IN CHINESE (简体中文).
    - year: The relevant year (e.g., "2023", "2021-2022"). If unknown, use "未知年份".
    - keywords: A string of comma-separated keywords (5-10 keywords). MUST BE IN CHINESE (简体中文).
    - standardized_filename: A standardized filename for this file. Format: "Type-Year-Subject-SpecificName(Note).Suffix". 
      "Type" should be one of [考试, 笔记, 作业, 幻灯片, 实验, 书籍, 代码, 其他].
      Ensure the filename is valid for a file system (no special chars like / : * ? " < > |). 
      Example: "考试-2022-线性代数-期末考试(A卷).pdf". MUST BE IN CHINESE (简体中文) where appropriate, except for specific codes.
    """

    user_prompt = f"""
    File Path: {file_path}
    File Name: {meta['FileName']}
    File Size: {meta['FileSize']} bytes
    Neighbor Files: {context}
    
    File Content Snippet (First 5000 chars):
    {content_to_analyze[:5000]}
    """
    
    logger.info(f"Analyzing {meta['FileName']} with AI1...")
    response_json = ai1.chat_completion([
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt}
    ], json_mode=True)
    
    result = {
        "description": "Analysis failed",
        "short_description": "Analysis failed",
        "subject": "未知",
        "year": "未知年份",
        "keywords": "",
        "standardized_filename": ""
    }
    
    if response_json:
        try:
            # Clean up potential markdown code blocks
            cleaned_json = response_json
            if "```" in cleaned_json:
                import re
                match = re.search(r"```(?:json)?(.*?)```", cleaned_json, re.DOTALL)
                if match:
                    cleaned_json = match.group(1).strip()
            
            result = json.loads(cleaned_json)
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON from AI1: {e}")
            logger.debug(f"Raw response: {response_json}")

    # Step 3: Get Embedding (using Description + Keywords + Subject)
    embedding_text = f"{result.get('subject', '')} {result.get('keywords', '')} {result.get('short_description', '')}"
    
    # Use embedding_client if provided (Remote mode), else use ai1 (Local mode or Default)
    # Actually, current logic attaches local_model to ai1. 
    # If embedding_client is set, it means Remote. 
    # If ai1 has local_model, it means Local.
    
    if embedding_client:
        embedding = embedding_client.get_embedding(embedding_text)
    else:
        embedding = ai1.get_embedding(embedding_text)
    
    return result, embedding

def main():
    parser = argparse.ArgumentParser(description="AI File Indexer")
    parser.add_argument("--input", required=True, help="Input raw files directory")
    parser.add_argument("--output", required=True, help="Output hashed files directory")
    parser.add_argument("--db", required=True, help="Path to SQLite DB")
    
    parser.add_argument("--ai1-host", required=True, help="LLM API Host")
    parser.add_argument("--ai1-key", required=True, help="LLM API Key")
    parser.add_argument("--ai1-model", default="gpt-4o", help="LLM Model Name")
    
    parser.add_argument("--ai2-host", required=True, help="OCR/Vision API Host")
    parser.add_argument("--ai2-key", required=True, help="OCR/Vision API Key")
    parser.add_argument("--ai2-model", default="gpt-4o", help="OCR/Vision Model Name")
    
    parser.add_argument("--embedding-local", action='store_true', help="Use local sentence-transformers for embedding instead of API")
    parser.add_argument("--embedding-host", default=None, help="Embedding API Host (if not local)")
    parser.add_argument("--embedding-key", default=None, help="Embedding API Key (if not local)")
    parser.add_argument("--embedding-model", default="all-MiniLM-L6-v2", help="Embedding model name")
    
    parser.add_argument("--force", action='store_true', help="Force re-indexing even if file exists in DB")

    args = parser.parse_args()

    # Init DB
    init_db(args.db)
    
    # Init Session
    session = IndexSession(args.db)
    
    # Init Clients
    ai1 = AIClient(args.ai1_host, args.ai1_key, args.ai1_model)
    ai2 = AIClient(args.ai2_host, args.ai2_key, args.ai2_model)
    
    # Init Embedding Client
    embedding_client = None
    if args.embedding_local:
        try:
            from sentence_transformers import SentenceTransformer
            logger.info(f"Loading local embedding model: {args.embedding_model}")
            # Attach local model to ai1 for legacy method call or create new structure
            # Let's keep ai1 as the embedder container for simplicity in analyze_content
            ai1.local_model = SentenceTransformer(args.embedding_model)
        except ImportError:
            logger.error("sentence-transformers not installed. Please run: pip install sentence-transformers")
            sys.exit(1)
    else:
        # Remote Embedding
        # If specific embedding host provided, use it. Else fall back to ai1 defaults
        emb_host = args.embedding_host or args.ai1_host
        emb_key = args.embedding_key or args.ai1_key
        emb_model = args.embedding_model
        
        # We need a way to tell ai1 to use these specific creds for embedding
        # Let's add an explicit embed_client to ai1 or pass it to analyze_content
        embedding_client = AIClient(emb_host, emb_key, emb_model)

    # Walk
    logger.info(f"Scanning directory: {args.input}")
    try:
        for root, dirs, files in os.walk(args.input):
            logger.info(f"Scanning subdirectory: {root} - Files found: {len(files)}")
            # Prepare context (neighbors)
            neighbors = [f for f in files if not f.startswith('.')]
            neighbor_context = ", ".join(neighbors[:10]) # First 10 neighbors names
            
            for file_name in files:
                if file_name.startswith('.'): continue
                
                file_path = os.path.join(root, file_name)
                
                # Check if likely already processed? (Optional optimization, skipping for now to enforce re-index)
                
                # Stats
                stat = os.stat(file_path)
                ext = os.path.splitext(file_name)[1].lower()
                
                
                import hashlib
                sha256 = hashlib.sha256()
                with open(file_path, "rb") as f:
                    while chunk := f.read(65536):
                        sha256.update(chunk)
                file_hash = sha256.hexdigest()
                
                # Check Cache
                if not args.force and check_if_indexed(args.db, file_hash):
                    logger.info(f"Skipping cached file: {file_name} (Hash: {file_hash[:8]}...)")
                    session.log_skip(ext[1:] if ext else "unknown")
                    # Ensure physical file exists in output even if DB has it
                    dest_path = os.path.join(args.output, file_hash)

                    if not os.path.exists(dest_path):
                        import shutil
                        try:
                            shutil.copy2(file_path, dest_path)
                            logger.info(f"  -> Restored physical file for {file_name}")
                        except Exception as e:
                            logger.error(f"  -> Failed to restore physical file: {e}")
                    continue
                
                # Prepare Meta
                meta = {
                    "FileHash": file_hash,
                    "FileName": os.path.splitext(file_name)[0], # Logic to avoid dupes needed?
                                                                 # The Java code handles dupes by just adding. 
                                                                 # The Python UpdateDB handled dupes by appending folder. 
                                                                 # We should keep simple for now or copy logic.
                    "FileType": ext[1:] if ext else "",
                    "FileType_Ext": ext, # Internal use
                    "FileSize": stat.st_size,
                    "Time": int(stat.st_mtime),
                    "Description": "",
                    "ShortDescription": "",
                    "Subject": "",
                    "Year": "",
                    "Keywords": "",
                    "Embedding": []
                }

                
                if not args.force and check_if_indexed(args.db, file_hash):
                     logger.info(f"Already indexed (skipping): {file_name}")
                     session.log_skip(meta['FileType'])
                     continue 

                
                # 2. Extract & Analyze
                text_content = extract_text_from_file(file_path, ext)
                
                # AI Analysis
                analysis_result, embedding = analyze_content(ai1, ai2, file_path, meta, neighbor_context, text_content, embedding_client)
                
                # Check for failure
                if not embedding or analysis_result.get('description') == "Analysis failed":
                    logger.error(f"Skipping DB save for {file_name} due to analysis failure.")
                    session.log_fail(meta['FileType'])
                    continue
                    
                session.log_success(meta['FileType'])
                
                meta.update({
                    "Description": analysis_result.get('description'),
                    "ShortDescription": analysis_result.get('short_description'),
                    "Subject": analysis_result.get('subject'),
                    "Year": analysis_result.get('year'),
                    "Keywords": analysis_result.get('keywords'),
                    "Embedding": embedding
                })

                # Handle standardized filename renaming (Metadata update)
                std_filename = analysis_result.get('standardized_filename')
                if std_filename and len(std_filename) > 3:
                    # Basic sanitation since AI might return invalid chars
                    import re
                    std_filename = re.sub(r'[\\/*?:"<>|]', "", std_filename)
                    
                    # Split name and ext
                    if '.' in std_filename:
                        new_fname, new_ext = std_filename.rsplit('.', 1)
                        if new_fname.strip():
                            meta['FileName'] = new_fname.strip()
                            logger.info(f"  -> Renamed to: {meta['FileName']}.{meta['FileType']}")
                    else:
                        if std_filename.strip():
                            meta['FileName'] = std_filename.strip()
                            logger.info(f"  -> Renamed to: {meta['FileName']} (ext kept)")
                
                # 3. Save
                # Move file
                dest_path = os.path.join(args.output, file_hash)
                if not os.path.exists(dest_path):
                    import shutil
                    shutil.copy2(file_path, dest_path) # Copy instead of move for safety in testing
                
                # Save to DB
                # We need to handle Unique FileName constraint manually if it fails?
                # For now, let's try/except
                try:
                    save_to_db(args.db, meta)
                    logger.info(f"Indexed: {meta['FileName']}.{meta['FileType']}")
                except sqlite3.IntegrityError:
                    # Handle duplicate name by appending hash segment
                    meta['FileName'] = f"{meta['FileName']}_{file_hash[:6]}"
                    save_to_db(args.db, meta)
                    logger.info(f"Indexed (Renamed for Unique): {meta['FileName']}.{meta['FileType']}")
                
    finally:
        session.finish()
        logger.info(f"Index Session Finished. Success: {session.success}, Failed: {session.failed}, Skipped: {session.skipped}")
                
if __name__ == "__main__":
    main()
